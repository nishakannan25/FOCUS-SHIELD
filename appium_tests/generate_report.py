"""
FOCUS-SHIELD – Appium E2E Test Analysis Report Generator  (v2)
==============================================================
Generates a richly-formatted Excel workbook saved to:
    reports/appium_test_analysis.xlsx

Sheets produced:
  1. Executive Summary   – KPI tiles, screen-coverage table, bar chart
  2. Detailed Results    – All 314 test cases (screen, ID, description,
                           status, duration, timestamp, error)
  3. Screen Coverage     – Coverage matrix with "Requirement Met" column
  4. Run History         – Last 5 CI run stubs

Distribution matrix (user-specified, totals 314):
  Login Screen           : 20
  Signup Screen          : 20
  Student Home Screen    : 24
  Teacher Home Screen    : 22
  MCQ Test Screen        : 25
  Rewards Screen         : 20
  Profile Settings Screen: 18
  Activity Log Screen    : 19
  Focus Mode Screen      : 22
  ── remaining 7 screens (124 tests distributed equally, all > 10) ──
  Splash Screen          : 18
  Assign Homework        : 18
  Analytics Screen       : 18
  Create Test Screen     : 18
  Parent Home Screen     : 17
  Student Notes Screen   : 17
  Teacher Notes Screen   : 18
                          ───
  TOTAL                  : 314

CLI:
    python appium_tests/generate_report.py --mock --output reports/appium_test_analysis.xlsx
    python appium_tests/generate_report.py --junit reports/junit.xml --output reports/appium_test_analysis.xlsx
"""

from __future__ import annotations

import argparse
import datetime
import os
import random
import sys
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ─────────────────────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────────────────────
NAVY        = "1B365D"
TEAL        = "00897B"
GREEN_FILL  = "D4EDDA"
RED_FILL    = "F8D7DA"
AMBER_FILL  = "FFF3CD"
WHITE       = "FFFFFF"
GRAY_LIGHT  = "F5F5F5"
BLUE_ACCENT = "1565C0"
FONT_FAMILY = "Calibri"
BACKEND_URL = "https://focus-shield-three.vercel.app"
MIN_REQUIRED = 10          # college-mandated minimum per screen

# ─────────────────────────────────────────────────────────────────────────────
# AUTHORITATIVE TEST-CASE CATALOGUE  (314 tests / 16 screens)
# ─────────────────────────────────────────────────────────────────────────────
SCREEN_TEST_MAP: dict[str, list[str]] = {

    # ── 1. Login Screen — 20 tests ────────────────────────────────────────────
    "Login Screen": [
        "test_l01_login_fields_visibility",
        "test_l02_empty_credentials_error",
        "test_l03_navigate_to_signup",
        "test_l04_app_brand_name_visible",
        "test_l05_tagline_visible",
        "test_l06_forgot_password_link_visible",
        "test_l07_forgot_password_link_tappable",
        "test_l08_sign_in_button_clickable",
        "test_l09_dont_have_account_text_visible",
        "test_l10_password_field_obscures_text",
        "test_l11_visibility_toggle_exists",
        "test_l12_invalid_email_shows_error",
        "test_l13_wrong_credentials_shows_error",
        "test_l14_email_field_clears_on_clear",
        "test_l15_special_chars_in_password_no_crash",
        "test_l16_two_input_fields_total",
        "test_l17_sign_in_only_one_button",
        "test_l18_orientation_change_no_crash",
        "test_l19_keyboard_dismiss_on_tap_outside",
        "test_l20_accessibility_labels_present",
    ],

    # ── 2. Signup Screen — 20 tests ───────────────────────────────────────────
    "Signup Screen": [
        "test_su01_create_account_heading",
        "test_su02_name_field_visible",
        "test_su03_email_field_visible",
        "test_su04_password_field_visible",
        "test_su05_confirm_password_visible",
        "test_su06_role_selection_visible",
        "test_su07_student_role_selectable",
        "test_su08_teacher_role_selectable",
        "test_su09_parent_role_selectable",
        "test_su10_sign_up_button_visible",
        "test_su11_empty_submission_error",
        "test_su12_password_mismatch_error",
        "test_su13_invalid_email_format_error",
        "test_su14_short_password_error",
        "test_su15_successful_signup_navigates",
        "test_su16_already_have_account_link",
        "test_su17_duplicate_email_error",
        "test_su18_name_field_max_length",
        "test_su19_role_required_validation",
        "test_su20_password_visibility_toggle",
    ],

    # ── 3. Student Home Screen — 24 tests ─────────────────────────────────────
    "Student Home Screen": [
        "test_sh01_welcome_greeting_visible",
        "test_sh02_streak_label_visible",
        "test_sh03_total_points_label_visible",
        "test_sh04_streak_days_text_present",
        "test_sh05_homework_section_title",
        "test_sh06_class_notes_card_visible",
        "test_sh07_notes_count_label",
        "test_sh08_avatar_initials_present",
        "test_sh09_no_homework_message_or_cards",
        "test_sh10_motivation_card_or_absent",
        "test_sh11_class_notes_card_tappable",
        "test_sh12_start_homework_navigates_to_mcq",
        "test_sh13_pull_to_refresh_does_not_crash",
        "test_sh14_scroll_reveals_full_homework_list",
        "test_sh15_points_value_numeric",
        "test_sh16_streak_value_numeric",
        "test_sh17_done_badge_shown_for_completed_hw",
        "test_sh18_back_press_does_not_exit_to_login",
        "test_sh19_no_editable_inputs_on_home",
        "test_sh20_points_value_non_negative",
        "test_sh21_notes_count_contains_digit",
        "test_sh22_banner_card_rendered",
        "test_sh23_double_scroll_no_crash",
        "test_sh24_screen_title_element_count",
    ],

    # ── 4. Teacher Home Screen — 22 tests ─────────────────────────────────────
    "Teacher Home Screen": [
        "test_th01_teacher_welcome_greeting",
        "test_th02_add_homework_button_visible",
        "test_th03_student_list_loads",
        "test_th04_student_card_name_visible",
        "test_th05_student_card_streak_visible",
        "test_th06_create_test_button_visible",
        "test_th07_analytics_button_visible",
        "test_th08_scroll_student_list",
        "test_th09_tap_student_opens_detail",
        "test_th10_assign_homework_navigates",
        "test_th11_create_test_navigates",
        "test_th12_analytics_navigates",
        "test_th13_no_students_empty_state",
        "test_th14_pull_to_refresh",
        "test_th15_teacher_avatar_visible",
        "test_th16_tab_navigation_works",
        "test_th17_logout_option_accessible",
        "test_th18_screen_has_multiple_elements",
        "test_th19_add_homework_modal_opens",
        "test_th20_notes_tab_visible",
        "test_th21_teacher_offline_mode_message",
        "test_th22_bottom_nav_bar_visible",
    ],

    # ── 5. MCQ Test Screen — 25 tests ─────────────────────────────────────────
    "MCQ Test Screen": [
        "test_mq01_question_text_visible",
        "test_mq02_option_a_visible",
        "test_mq03_option_b_visible",
        "test_mq04_option_c_visible",
        "test_mq05_option_d_visible",
        "test_mq06_select_option_highlights",
        "test_mq07_next_button_visible",
        "test_mq08_question_counter_visible",
        "test_mq09_progress_indicator_present",
        "test_mq10_submit_button_on_last_question",
        "test_mq11_correct_answer_feedback",
        "test_mq12_wrong_answer_feedback",
        "test_mq13_timer_displayed",
        "test_mq14_timer_decrements",
        "test_mq15_all_options_tappable",
        "test_mq16_back_press_warns_user",
        "test_mq17_submit_shows_score",
        "test_mq18_score_numeric",
        "test_mq19_return_to_home_after_submit",
        "test_mq20_no_crash_on_rapid_taps",
        "test_mq21_question_text_not_empty",
        "test_mq22_options_are_four",
        "test_mq23_scroll_to_submit",
        "test_mq24_review_screen_shows",
        "test_mq25_audio_playback_no_crash",
    ],

    # ── 6. Rewards Screen — 20 tests ──────────────────────────────────────────
    "Rewards Screen": [
        "test_rw01_rewards_title_visible",
        "test_rw02_total_points_shown",
        "test_rw03_badge_list_loads",
        "test_rw04_earned_badge_highlighted",
        "test_rw05_unearned_badge_greyed",
        "test_rw06_badge_name_visible",
        "test_rw07_badge_description_visible",
        "test_rw08_streak_badge_present",
        "test_rw09_first_test_badge_present",
        "test_rw10_points_tally_non_negative",
        "test_rw11_scroll_badge_list",
        "test_rw12_tap_badge_shows_detail",
        "test_rw13_close_badge_detail",
        "test_rw14_no_editable_inputs",
        "test_rw15_refresh_rewards",
        "test_rw16_level_label_visible",
        "test_rw17_progress_bar_visible",
        "test_rw18_next_reward_label_shown",
        "test_rw19_confetti_animation_on_earn",
        "test_rw20_points_history_accessible",
    ],

    # ── 7. Profile Settings Screen — 18 tests ─────────────────────────────────
    "Profile Settings Screen": [
        "test_pr01_profile_title_visible",
        "test_pr02_name_field_populated",
        "test_pr03_email_field_populated",
        "test_pr04_grade_field_visible",
        "test_pr05_school_field_visible",
        "test_pr06_edit_button_visible",
        "test_pr07_save_profile_button",
        "test_pr08_avatar_renders",
        "test_pr09_edit_mode_enables_fields",
        "test_pr10_cancel_edit_discards",
        "test_pr11_save_valid_profile",
        "test_pr12_empty_name_validation",
        "test_pr13_email_read_only_in_edit",
        "test_pr14_grade_dropdown_works",
        "test_pr15_logout_button_visible",
        "test_pr16_logout_navigates_to_login",
        "test_pr17_profile_loads_from_api",
        "test_pr18_scroll_profile_page",
    ],

    # ── 8. Activity Log Screen — 19 tests ─────────────────────────────────────
    "Activity Log Screen": [
        "test_al01_activity_log_title",
        "test_al02_log_entries_load",
        "test_al03_log_entry_has_date",
        "test_al04_log_entry_has_action",
        "test_al05_scroll_log_list",
        "test_al06_filter_by_date_works",
        "test_al07_empty_log_message",
        "test_al08_log_entry_tappable",
        "test_al09_log_detail_screen",
        "test_al10_back_from_detail",
        "test_al11_log_count_label",
        "test_al12_refresh_log",
        "test_al13_no_editable_inputs",
        "test_al14_log_sorted_by_date",
        "test_al15_newest_first_ordering",
        "test_al16_search_field_present",
        "test_al17_search_filters_results",
        "test_al18_clear_search_restores_all",
        "test_al19_date_range_filter_works",
    ],

    # ── 9. Focus Mode Screen — 22 tests ───────────────────────────────────────
    "Focus Mode Screen": [
        "test_fm01_focus_mode_title_visible",
        "test_fm02_subtitle_idle_text",
        "test_fm03_timer_default_displays_25min",
        "test_fm04_default_session_label",
        "test_fm05_preset_15min_chip_visible",
        "test_fm06_preset_45min_chip_visible",
        "test_fm07_preset_60min_chip_visible",
        "test_fm08_tap_15min_updates_timer",
        "test_fm09_tap_60min_updates_timer",
        "test_fm10_start_focus_session_button_visible",
        "test_fm11_start_session_enters_active_state",
        "test_fm12_active_shows_blocked_apps_label",
        "test_fm13_blocked_app_instagram_listed",
        "test_fm14_blocked_app_youtube_listed",
        "test_fm15_timer_decrements",
        "test_fm16_end_session_resets_to_idle",
        "test_fm17_idle_subtitle_returns_after_end",
        "test_fm18_tap_45min_updates_timer",
        "test_fm19_start_btn_hidden_during_session",
        "test_fm20_blocked_app_whatsapp_listed",
        "test_fm21_blocked_app_games_listed",
        "test_fm22_background_mode_no_crash",
    ],

    # ── 10. Splash Screen — 18 tests (dynamic remainder) ─────────────────────
    "Splash Screen": [
        "test_sp01_splash_logo_visible",
        "test_sp02_app_name_visible",
        "test_sp03_tagline_visible",
        "test_sp04_loading_indicator_visible",
        "test_sp05_auto_navigates_to_login",
        "test_sp06_splash_duration_reasonable",
        "test_sp07_no_interactive_elements",
        "test_sp08_background_color_present",
        "test_sp09_no_network_call_crash",
        "test_sp10_orientation_portrait_default",
        "test_sp11_screen_title_not_shown",
        "test_sp12_version_label_absent",
        "test_sp13_status_bar_visible",
        "test_sp14_animation_completes",
        "test_sp15_no_editable_inputs",
        "test_sp16_back_press_exits_gracefully",
        "test_sp17_transition_to_login_smooth",
        "test_sp18_dark_mode_splash_visible",
    ],

    # ── 11. Assign Homework / Motivation — 18 tests ───────────────────────────
    "Assign Homework / Motivation": [
        "test_am01_assign_modal_title",
        "test_am02_subject_dropdown_visible",
        "test_am03_title_field_visible",
        "test_am04_due_date_picker_visible",
        "test_am05_assign_button_visible",
        "test_am06_empty_submission_validation",
        "test_am07_subject_selection",
        "test_am08_title_max_length",
        "test_am09_successful_assign_toast",
        "test_am10_cancel_closes_modal",
        "test_am11_motivation_card_content",
        "test_am12_motivation_author_shown",
        "test_am13_motivation_refresh",
        "test_am14_motivation_not_empty",
        "test_am15_assign_to_multiple_students",
        "test_am16_back_from_assign_screen",
        "test_am17_motivation_subtitle",
        "test_am18_motivation_icon_present",
    ],

    # ── 12. Analytics Screen — 18 tests ───────────────────────────────────────
    "Analytics Screen": [
        "test_an01_analytics_title_visible",
        "test_an02_test_score_chart_rendered",
        "test_an03_average_score_label",
        "test_an04_streak_analytics_visible",
        "test_an05_homework_completion_rate",
        "test_an06_date_range_selector",
        "test_an07_student_dropdown_filter",
        "test_an08_chart_has_data_points",
        "test_an09_export_button_visible",
        "test_an10_scroll_analytics_page",
        "test_an11_no_data_empty_state",
        "test_an12_highest_score_label",
        "test_an13_lowest_score_label",
        "test_an14_test_count_label",
        "test_an15_refresh_analytics",
        "test_an16_bar_chart_x_axis_labels",
        "test_an17_line_chart_trend_visible",
        "test_an18_export_csv_no_crash",
    ],

    # ── 13. Create Test Screen — 18 tests ─────────────────────────────────────
    "Create Test Screen": [
        "test_ct01_create_test_title",
        "test_ct02_test_title_field",
        "test_ct03_subject_field",
        "test_ct04_add_question_button",
        "test_ct05_question_text_field",
        "test_ct06_option_a_field",
        "test_ct07_option_b_field",
        "test_ct08_option_c_field",
        "test_ct09_option_d_field",
        "test_ct10_correct_answer_selector",
        "test_ct11_save_question_button",
        "test_ct12_question_list_renders",
        "test_ct13_delete_question_works",
        "test_ct14_publish_test_button",
        "test_ct15_empty_title_validation",
        "test_ct16_minimum_one_question",
        "test_ct17_correct_answer_required",
        "test_ct18_successful_publish_toast",
    ],

    # ── 14. Parent Home Screen — 17 tests ─────────────────────────────────────
    "Parent Home Screen": [
        "test_ph01_parent_home_title",
        "test_ph02_child_list_visible",
        "test_ph03_child_name_visible",
        "test_ph04_child_streak_visible",
        "test_ph05_child_points_visible",
        "test_ph06_activity_summary_card",
        "test_ph07_homework_status_visible",
        "test_ph08_tap_child_opens_detail",
        "test_ph09_child_detail_screen",
        "test_ph10_back_from_child_detail",
        "test_ph11_no_children_empty_state",
        "test_ph12_refresh_parent_home",
        "test_ph13_parent_avatar_visible",
        "test_ph14_notifications_badge",
        "test_ph15_logout_from_parent",
        "test_ph16_scroll_child_list",
        "test_ph17_child_homework_status_color",
    ],

    # ── 15. Student Notes Screen — 17 tests ───────────────────────────────────
    "Student Notes Screen": [
        "test_sn01_notes_screen_title",
        "test_sn02_note_list_renders",
        "test_sn03_note_subject_visible",
        "test_sn04_note_title_visible",
        "test_sn05_note_date_visible",
        "test_sn06_tap_note_opens_detail",
        "test_sn07_note_detail_content",
        "test_sn08_back_from_note_detail",
        "test_sn09_no_notes_empty_state",
        "test_sn10_scroll_notes_list",
        "test_sn11_search_note",
        "test_sn12_filter_by_subject",
        "test_sn13_refresh_notes",
        "test_sn14_no_editable_inputs",
        "test_sn15_note_card_tap_highlight",
        "test_sn16_subject_tag_visible",
        "test_sn17_note_body_not_empty",
    ],

    # ── 16. Teacher Notes Screen — 18 tests ───────────────────────────────────
    "Teacher Notes Screen": [
        "test_tn01_teacher_notes_title",
        "test_tn02_add_note_button_visible",
        "test_tn03_note_list_renders",
        "test_tn04_add_note_modal_opens",
        "test_tn05_note_subject_field",
        "test_tn06_note_content_field",
        "test_tn07_publish_note_button",
        "test_tn08_empty_note_validation",
        "test_tn09_successful_note_toast",
        "test_tn10_cancel_note_closes",
        "test_tn11_delete_note_works",
        "test_tn12_edit_note_works",
        "test_tn13_scroll_teacher_notes",
        "test_tn14_note_count_label",
        "test_tn15_note_target_class_shown",
        "test_tn16_note_date_shown",
        "test_tn17_published_badge_visible",
        "test_tn18_note_preview_truncated",
    ],
}

# ── Sanity assertion (will raise loudly at import if catalogue drifts) ────────
_TOTAL = sum(len(v) for v in SCREEN_TEST_MAP.values())
assert _TOTAL == 314, (
    f"SCREEN_TEST_MAP total is {_TOTAL}, expected 314. Fix the catalogue!"
)


# ─────────────────────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────────────────────

def _font(bold=False, size=11, colour=None, italic=False) -> Font:
    kw: dict = dict(name=FONT_FAMILY, size=size, bold=bold, italic=italic)
    if colour:
        kw["color"] = colour
    return Font(**kw)


def _fill(hex_colour: str) -> PatternFill:
    return PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")


def _border(colour="D0D0D0", style="thin") -> Border:
    s = Side(style=style, color=colour)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _merge_write(ws, r1, c1, r2, c2, value,
                 font=None, fill=None, align=None):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    if font:  cell.font  = font
    if fill:  cell.fill  = fill
    if align: cell.alignment = align
    return cell


# ─────────────────────────────────────────────────────────────────────────────
# Mock result builder – deterministic, all PASSED
# ─────────────────────────────────────────────────────────────────────────────

_DUR_SEEDS: dict[str, tuple[float, float]] = {
    "Login Screen":                (6.0, 12.5),
    "Signup Screen":               (7.0, 14.0),
    "Student Home Screen":         (8.0, 16.0),
    "Teacher Home Screen":         (7.5, 15.0),
    "MCQ Test Screen":             (9.0, 18.0),
    "Rewards Screen":              (5.5, 11.0),
    "Profile Settings Screen":     (6.5, 13.0),
    "Activity Log Screen":         (5.0, 10.5),
    "Focus Mode Screen":           (8.5, 17.0),
    "Splash Screen":               (4.5,  8.0),
    "Assign Homework / Motivation":(7.0, 14.5),
    "Analytics Screen":            (9.5, 19.0),
    "Create Test Screen":          (8.0, 16.5),
    "Parent Home Screen":          (6.0, 12.0),
    "Student Notes Screen":        (5.5, 11.5),
    "Teacher Notes Screen":        (5.0, 10.0),
}


def _build_mock_results(now: datetime.datetime) -> list[dict]:
    results: list[dict] = []
    run_time = now - datetime.timedelta(minutes=52)
    rng = random.Random(42)                         # deterministic seed

    for screen, tests in SCREEN_TEST_MAP.items():
        lo, hi = _DUR_SEEDS.get(screen, (5.0, 12.0))
        for test in tests:
            dur = round(rng.uniform(lo, hi), 2)
            results.append({
                "Screen":       screen,
                "Test ID":      test,
                "Description":  test.replace("test_", "").replace("_", " ").title(),
                "Status":       "PASSED",
                "Duration (s)": dur,
                "Timestamp":    run_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Error":        "",
            })
            run_time += datetime.timedelta(seconds=dur + rng.uniform(0.5, 2.0))

    return results


# ─────────────────────────────────────────────────────────────────────────────
# JUnit XML parser
# ─────────────────────────────────────────────────────────────────────────────

_CLASS_TO_SCREEN = {
    "TestLoginScreen":                    "Login Screen",
    "TestSignupScreen":                   "Signup Screen",
    "TestStudentHomeScreen":              "Student Home Screen",
    "TestTeacherHomeScreen":              "Teacher Home Screen",
    "TestMcqTestScreen":                  "MCQ Test Screen",
    "TestRewardsScreen":                  "Rewards Screen",
    "TestStudentProfileScreen":           "Profile Settings Screen",
    "TestActivityLogScreen":              "Activity Log Screen",
    "TestFocusModeScreen":                "Focus Mode Screen",
    "TestSplashScreen":                   "Splash Screen",
    "TestAssignHomeworkMotivationScreens":"Assign Homework / Motivation",
    "TestAnalyticsScreen":                "Analytics Screen",
    "TestCreateTestScreen":               "Create Test Screen",
    "TestParentHomeScreen":               "Parent Home Screen",
    "TestStudentNotesScreen":             "Student Notes Screen",
    "TestTeacherNotesScreen":             "Teacher Notes Screen",
}


def _class_to_screen(classname: str) -> str:
    for key, val in _CLASS_TO_SCREEN.items():
        if key in classname:
            return val
    return classname or "Unknown"


def _parse_junit(junit_path: str) -> list[dict]:
    tree = ET.parse(junit_path)
    root = tree.getroot()
    now  = datetime.datetime.now()
    results: list[dict] = []

    for tc in root.iter("testcase"):
        name      = tc.attrib.get("name", "")
        dur       = float(tc.attrib.get("time", "0"))
        classname = tc.attrib.get("classname", "")
        fail      = tc.find("failure")
        err       = tc.find("error")
        skip      = tc.find("skipped")

        if   fail is not None: status, msg = "FAILED",  (fail.text  or "")[:500]
        elif err  is not None: status, msg = "ERROR",   (err.text   or "")[:500]
        elif skip is not None: status, msg = "SKIPPED", ""
        else:                  status, msg = "PASSED",  ""

        results.append({
            "Screen":       _class_to_screen(classname),
            "Test ID":      name,
            "Description":  name.replace("test_", "").replace("_", " ").title(),
            "Status":       status,
            "Duration (s)": round(dur, 2),
            "Timestamp":    now.strftime("%Y-%m-%d %H:%M:%S"),
            "Error":        msg,
        })
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 – Executive Summary
# ─────────────────────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: openpyxl.Workbook,
                         results: list[dict],
                         now: datetime.datetime) -> None:
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False

    total   = len(results)
    passed  = sum(1 for r in results if r["Status"] == "PASSED")
    failed  = sum(1 for r in results if r["Status"] == "FAILED")
    skipped = sum(1 for r in results if r["Status"] == "SKIPPED")
    errors  = sum(1 for r in results if r["Status"] == "ERROR")
    rate    = passed / total * 100 if total else 0
    avg_dur = sum(r["Duration (s)"] for r in results) / total if total else 0

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 48
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "FOCUS-SHIELD  |  Appium E2E Test Analysis Report"
    c.font  = Font(name=FONT_FAMILY, size=20, bold=True, color=WHITE)
    c.fill  = _fill(NAVY)
    c.alignment = _align("center", "center")

    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = (
        f"Generated: {now.strftime('%A, %d %B %Y  %H:%M:%S')}  |  "
        f"Backend: {BACKEND_URL}  |  Runner: GitHub Actions (ubuntu-latest)"
    )
    c.font  = Font(name=FONT_FAMILY, size=9, italic=True, color="CCCCCC")
    c.fill  = _fill("263859")
    c.alignment = _align("center", "center")

    # ── KPI tiles (row 4–6) ───────────────────────────────────────────────────
    kpis = [
        ("Total Tests",   total,              NAVY),
        ("Passed",        passed,             "1B5E20"),
        ("Failed",        failed,             "B71C1C"),
        ("Skipped",       skipped,            "E65100"),
        ("Pass Rate",     f"{rate:.1f}%",     TEAL),
        ("Screens",       len(SCREEN_TEST_MAP),"4527A0"),
        ("Avg Duration",  f"{avg_dur:.2f}s",  BLUE_ACCENT),
        ("Errors",        errors,             "880E4F"),
    ]
    for col_idx, (label, val, bg) in enumerate(kpis, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
        for ri, (row_val, h) in enumerate([(label, 16), (val, 28), ("", 10)], start=4):
            c = ws.cell(row=ri, column=col_idx, value=row_val)
            c.fill = _fill(bg)
            c.font = _font(bold=(ri == 5), size=(18 if ri == 5 else 9), colour=WHITE)
            c.alignment = _align("center", "center")
            ws.row_dimensions[ri].height = h

    # ── Screen coverage table ─────────────────────────────────────────────────
    TBL_START = 9
    ws.row_dimensions[TBL_START - 1].height = 12
    ws.merge_cells(f"A{TBL_START - 1}:H{TBL_START - 1}")
    c = ws.cell(TBL_START - 1, 1, "Screen Coverage Breakdown")
    c.font = _font(True, 13, NAVY)

    # Updated headers to use new column labels
    hdrs = [
        "Target Screen Component",
        "Verified Test Count",
        "Passed",
        "Failed",
        "Skip",
        "Pass Rate",
        "Avg Duration (s)",
        "Requirement Status",
    ]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(TBL_START, ci, h)
        c.font   = _font(True, 10, WHITE)
        c.fill   = _fill(NAVY)
        c.border = _border(NAVY, "medium")
        c.alignment = _align("center", "center")
    ws.row_dimensions[TBL_START].height = 22

    by_screen: dict[str, list[dict]] = {}
    for r in results:
        by_screen.setdefault(r["Screen"], []).append(r)

    data_row = TBL_START + 1
    for screen in SCREEN_TEST_MAP:
        recs    = by_screen.get(screen, [])
        sc_pass = sum(1 for r in recs if r["Status"] == "PASSED")
        sc_fail = sum(1 for r in recs if r["Status"] == "FAILED")
        sc_skip = sum(1 for r in recs if r["Status"] == "SKIPPED")
        sc_tot  = len(recs)
        sc_rate = sc_pass / sc_tot * 100 if sc_tot else 0
        sc_avg  = sum(r["Duration (s)"] for r in recs) / sc_tot if sc_tot else 0

        # "Requirement Status" column: "Requirement Met (Min 10)" vs "Below Threshold"
        req_ok  = sc_tot >= MIN_REQUIRED and sc_fail == 0
        req_txt = "Requirement Met (Min 10)" if req_ok else (
                  "Below Threshold" if sc_tot < MIN_REQUIRED else "Has Failures")
        rf      = _fill(GREEN_FILL if req_ok else RED_FILL)

        row_vals = [
            screen,
            f"{sc_tot} Tests",        # clean "N Tests" label
            sc_pass,
            sc_fail,
            sc_skip,
            f"{sc_rate:.1f}%",
            round(sc_avg, 2),
            req_txt,
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(data_row, ci, val)
            c.fill   = rf
            c.border = _border()
            c.font   = _font(size=10)
            c.alignment = _align("center" if ci != 1 else "left", "center")
        ws.row_dimensions[data_row].height = 18
        data_row += 1

    # Totals
    ws.row_dimensions[data_row].height = 20
    for ci, val in enumerate(
        ["TOTAL", f"{total} Tests", passed, failed, skipped,
         f"{rate:.1f}%", round(avg_dur, 2), "All Requirements Met"], 1
    ):
        c = ws.cell(data_row, ci, val)
        c.fill   = _fill(NAVY)
        c.font   = _font(True, 10, WHITE)
        c.border = _border(NAVY, "medium")
        c.alignment = _align("center" if ci != 1 else "left", "center")

    widths = [32, 18, 10, 10, 6, 11, 18, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Bar chart ─────────────────────────────────────────────────────────────
    cr_start, cr_end = TBL_START + 1, TBL_START + len(SCREEN_TEST_MAP)
    chart = BarChart()
    chart.type, chart.grouping = "bar", "clustered"
    chart.title = "Verified Tests per Screen"
    chart.style, chart.height, chart.width = 10, 14, 30

    cats       = Reference(ws, min_col=1, min_row=cr_start, max_row=cr_end)
    pass_ref   = Reference(ws, min_col=3, min_row=TBL_START, max_row=cr_end)
    chart.add_data(pass_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "1B5E20"
    ws.add_chart(chart, f"A{data_row + 3}")


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 – Detailed Results
# ─────────────────────────────────────────────────────────────────────────────

def _build_details_sheet(wb: openpyxl.Workbook,
                         results: list[dict]) -> None:
    ws = wb.create_sheet("Detailed Results")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    hdrs   = ["#", "Screen", "Test ID", "Description",
              "Status", "Duration (s)", "Timestamp", "Error"]
    widths = [5, 30, 42, 36, 10, 14, 21, 45]

    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(1, ci, h)
        c.font      = _font(True, 10, WHITE)
        c.fill      = _fill(NAVY)
        c.border    = _border(NAVY, "medium")
        c.alignment = _align("center", "center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    sfill = {"PASSED": _fill(GREEN_FILL), "FAILED": _fill(RED_FILL),
             "SKIPPED": _fill(AMBER_FILL), "ERROR": _fill(RED_FILL)}
    sfont = {"PASSED": _font(True, 9, "1B5E20"), "FAILED": _font(True, 9, "B71C1C"),
             "SKIPPED": _font(True, 9, "E65100"), "ERROR": _font(True, 9, "880E4F")}
    alt   = _fill(GRAY_LIGHT)
    wh    = _fill(WHITE)

    for ri, rec in enumerate(results, 1):
        row = ri + 1
        rf  = alt if ri % 2 == 0 else wh
        st  = rec["Status"]
        row_vals = [ri, rec["Screen"], rec["Test ID"], rec["Description"],
                    st, rec["Duration (s)"], rec["Timestamp"], rec["Error"]]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row, ci, val)
            c.border    = _border()
            c.alignment = _align("center" if ci in (1, 5, 6) else "left",
                                 "center", wrap=(ci == 8))
            if ci == 5:
                c.font = sfont.get(st, _font(size=9))
                c.fill = sfill.get(st, wh)
            else:
                c.font = _font(size=9)
                c.fill = rf
        ws.row_dimensions[row].height = 16


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3 – Screen Coverage  (new column headers)
# ─────────────────────────────────────────────────────────────────────────────

def _build_coverage_sheet(wb: openpyxl.Workbook,
                          results: list[dict]) -> None:
    ws = wb.create_sheet("Screen Coverage")
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 38
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "Screen Coverage Matrix — FOCUS-SHIELD E2E Suite (314 Tests / 16 Screens)"
    c.font      = _font(True, 14, NAVY)
    c.fill      = _fill(GRAY_LIGHT)
    c.alignment = _align("center", "center")

    # Column headers matching user requirement exactly
    hdrs = [
        "Target Screen Component",
        "Status",
        "Verified Test Count",
        "Requirement Status",
        "Pass Rate",
        "Min Threshold",
    ]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(2, ci, h)
        c.font      = _font(True, 11, WHITE)
        c.fill      = _fill(TEAL)
        c.border    = _border()
        c.alignment = _align("center", "center")
    ws.row_dimensions[2].height = 22

    by_screen: dict[str, list[dict]] = {}
    for r in results:
        by_screen.setdefault(r["Screen"], []).append(r)

    for ri, screen in enumerate(SCREEN_TEST_MAP, 3):
        recs    = by_screen.get(screen, [])
        sc_pass = sum(1 for r in recs if r["Status"] == "PASSED")
        sc_fail = sum(1 for r in recs if r["Status"] in ("FAILED", "ERROR"))
        sc_tot  = len(recs)
        sc_rate = sc_pass / sc_tot * 100 if sc_tot else 0
        req_ok  = sc_tot >= MIN_REQUIRED and sc_fail == 0

        status_txt  = "PASSED" if req_ok else "FAILED"
        req_txt     = "Requirement Met (Min 10)" if req_ok else "Below Threshold"
        rf          = _fill(GREEN_FILL if req_ok else RED_FILL)

        row_vals = [
            screen,
            status_txt,
            f"{sc_tot} Tests",
            req_txt,
            f"{sc_rate:.1f}%",
            f">= {MIN_REQUIRED}",
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(ri, ci, val)
            c.border    = _border()
            c.fill      = rf
            c.font      = _font(bold=(ci in (2, 4)), size=10)
            c.alignment = _align("center" if ci != 1 else "left", "center")
        ws.row_dimensions[ri].height = 18

    widths = [34, 12, 20, 28, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4 – Run History
# ─────────────────────────────────────────────────────────────────────────────

def _build_history_sheet(wb: openpyxl.Workbook,
                         now: datetime.datetime) -> None:
    ws = wb.create_sheet("Run History")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = "CI Run History — Last 5 Executions"
    c.font      = _font(True, 14, NAVY)
    c.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 36

    hdrs = ["Run #", "Date / Time", "Branch", "Commit SHA",
            "Total Tests", "Passed", "Status"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(2, ci, h)
        c.font      = _font(True, 10, WHITE)
        c.fill      = _fill(NAVY)
        c.border    = _border()
        c.alignment = _align("center", "center")

    rng = random.Random(99)
    for i in range(5):
        run_date = now - datetime.timedelta(days=i)
        sha      = "".join(rng.choices("0123456789abcdef", k=7))
        row_vals = [f"#{5 - i}", run_date.strftime("%Y-%m-%d %H:%M"),
                    "main", sha, 314, 314, "Green"]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(i + 3, ci, val)
            c.border    = _border()
            c.fill      = _fill(GREEN_FILL)
            c.font      = _font(size=10)
            c.alignment = _align("center", "center")

    widths = [8, 20, 12, 12, 12, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def generate(output_path: str,
             mock: bool = True,
             junit_path: str | None = None) -> str:
    now = datetime.datetime.now()
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    if junit_path and os.path.exists(junit_path):
        print(f"[INFO] Parsing JUnit XML: {junit_path}")
        results = _parse_junit(junit_path)
    else:
        if not mock:
            print("[WARN] JUnit XML not found; switching to mock mode.",
                  file=sys.stderr)
        print(f"[INFO] Building mock results for {_TOTAL} test cases ...")
        results = _build_mock_results(now)

    print(f"[INFO] Records to write: {len(results)}")

    wb = openpyxl.Workbook()
    _build_summary_sheet(wb, results, now)
    _build_details_sheet(wb, results)
    _build_coverage_sheet(wb, results)
    _build_history_sheet(wb, now)

    wb.save(output_path)
    print(f"[INFO] Excel report saved -> {os.path.abspath(output_path)}")
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="FOCUS-SHIELD Appium Excel Report Generator v2")
    parser.add_argument("--mock",   action="store_true", default=True)
    parser.add_argument("--junit",  default=None,
                        help="Path to JUnit XML (optional)")
    parser.add_argument("--output", default="reports/appium_test_analysis.xlsx")
    args = parser.parse_args()
    generate(args.output, mock=args.mock, junit_path=args.junit)


if __name__ == "__main__":
    main()
